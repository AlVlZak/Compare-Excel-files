# This is a script to compare two Excel files.
__version__ = "1.0"
import openpyxl as op
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import os

import argparse
import tkinter as tk
from tkinter import filedialog, messagebox
import time

import main as m


def find_difference_whole(wb1_path, wb2_path, output_path, value_or_formula="value", max_rows=1048576, max_columns=16384):
    wb1_name = os.path.basename(wb1_path)
    wb2_name = os.path.basename(wb2_path)
    separator = "--->"

    wb1, wb2, wb1_v, wb2_v, wb_output, formula = set_value_or_formula(wb1_path, wb2_path, value_or_formula)
    sheets_to_check, sheets_without_pair = find_sheet_to_compare(wb1, wb2, wb1_name, wb2_name)
    preparation_output_file_whole(wb_output, sheets_without_pair, value_or_formula)

    len_of_sheets_to_check = len(sheets_to_check)
    for k, sheet in enumerate(sheets_to_check, 1):
        max_row = min(max_rows, max(wb1[sheet].max_row, wb2[sheet].max_row))
        max_column = min(max_columns, max(wb1[sheet].max_column, wb2[sheet].max_column))

        df1_to_check, df2_to_check, wb_res, sheet = define_sheet(wb1, wb2, wb_output, sheet, sheet)

        print_status(sheet, k, max_row, max_column, len_of_sheets_to_check)

        rows, cols, df1_to_check, df2_to_check = compare_sheet(df1_to_check, df2_to_check, max_row, max_column)
        if formula:
            df1_to_check, df2_to_check = replace_formulas_by_values(wb1_v[sheet], wb2_v[sheet], max_row, max_column)
        rows = filling_output_df(df1_to_check, df2_to_check, rows, cols, separator)
        add_comparing_values_to_output(wb_res, rows, separator)
    print("\nCreating output file")
    output_name = os.path.join(output_path, "compare_excels.xlsx")
    wb_output.save(output_name)
    if formula:
        wb1_v.close()
        wb2_v.close()
    wb1.close()
    wb2.close()


def find_difference_selected(wb1_path, wb2_path, output_path, selected_sheets, value_or_formula="value", max_rows=1048576, max_columns=16384):
    wb1_name = os.path.basename(wb1_path)
    wb2_name = os.path.basename(wb2_path)
    separator = "--->"

    wb1, wb2, wb1_v, wb2_v, wb_output, formula = set_value_or_formula(wb1_path, wb2_path, value_or_formula)
    preparation_output_file_selected(wb_output, wb1_name, wb2_name, selected_sheets, value_or_formula)
    len_of_sheets_to_check = len(selected_sheets)
    for k, sheet in enumerate(selected_sheets, 1):
        df1_to_check, df2_to_check, wb_res, sheet_output = define_sheet(wb1, wb2, wb_output, sheet[0], sheet[1])
        max_row = min(max_rows, max(wb1[sheet[0]].max_row, wb2[sheet[1]].max_row))
        max_column = min(max_columns, max(wb1[sheet[0]].max_column, wb2[sheet[1]].max_column))

        print_status(sheet, k, max_row, max_column, len_of_sheets_to_check)

        rows, cols, df1_to_check, df2_to_check = compare_sheet(df1_to_check, df2_to_check, max_row, max_column)
        if formula:
            df1_to_check, df2_to_check = replace_formulas_by_values(wb1_v[sheet[0]], wb2_v[sheet[1]], max_row, max_column)

        rows = filling_output_df(df1_to_check, df2_to_check, rows, cols, separator)
        add_comparing_values_to_output(wb_res, rows, separator)
    print("\nCreating output file")
    output_name = os.path.join(output_path, "Find_diff_results.xlsx")
    wb_output.save(output_name)
    if formula:
        wb1_v.close()
        wb2_v.close()
    wb1.close()
    wb2.close()


def print_status(sheet, k, max_row, max_column, len_of_sheets_to_check):
    if not isinstance(sheet, list):
        print("Working with sheet \"{}\", max row is {}, max column is {},  progress {}/{}".format(sheet, max_row,
                                                                                                   max_column, k,
                                                                                                   len_of_sheets_to_check))
    else:
        print("Sheet \"{}\" is comparing with sheet \"{}\", max row is {}, max column is {},  progress {}/{}".format(sheet[0], sheet[1], max_row,
                                                                                                   max_column, k,
                                                                                                   len_of_sheets_to_check))
    # print("Working with sheet \"{}\",  progress {}/{}".format(sheet, k, len_of_sheets_to_check))


def filling_output_df(df1_to_check, df2_to_check, rows, cols, separator):
    for item in zip(rows, cols):
        output = '{} {} {}'.format(df1_to_check.iloc[item[0], item[1]], separator, df2_to_check.iloc[item[0], item[1]])
        empty_cell = [f"nan {separator} nan", f"None {separator} ", f"nan {separator} ", f" {separator} None", f" {separator} nan"]
        if output in empty_cell:
            output = ""
        output = output.replace("None", "")
        output = output.replace("nan", "")
        df1_to_check.iloc[item[0], item[1]] = output
    rows = dataframe_to_rows(df1_to_check, index=False, header=False)
    return rows


def add_comparing_values_to_output(wb_res, rows, separator):
    color_list = []
    color_tab = True
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            wb_res.cell(row=r_idx, column=c_idx, value=value)
            if separator in str(value):
                color_list.append([r_idx, c_idx])
    for c in color_list:
        wb_res.cell(row=c[0], column=c[1]).fill = op.styles.PatternFill("solid", start_color="FF0000")
        if color_tab:
            wb_res.sheet_properties.tabColor = "FF0000"
            color_tab = False


def replace_formulas_by_values(wb1_v, wb2_v, max_row, max_column):
    df1_to_check = pd.DataFrame(wb1_v.values)
    df2_to_check = pd.DataFrame(wb2_v.values)
    df1_to_check, df2_to_check = check_shape_and_set_to_one_shape(df1_to_check, df2_to_check, max_row, max_column)
    return df1_to_check, df2_to_check


def define_sheet(wb1, wb2, wb_output, sheet1, sheet2=""):
    if (sheet1 == sheet2) or (sheet2 == ""):
        sheet = sheet1
        sheet2 = sheet1
    else:
        sheet = sheet1[:15] + "&" + sheet2[:15]
    wb_output.create_sheet(sheet)
    wb1_to_check = wb1[sheet1]
    wb2_to_check = wb2[sheet2]
    wb_res = wb_output[sheet]
    df1_to_check = pd.DataFrame(wb1_to_check.values)
    df2_to_check = pd.DataFrame(wb2_to_check.values)
    return df1_to_check, df2_to_check, wb_res, sheet


def compare_sheet(df1_to_check, df2_to_check, max_row, max_column):
    df1_to_check, df2_to_check = check_shape_and_set_to_one_shape(df1_to_check, df2_to_check, max_row, max_column)
    comparison_values = df1_to_check.values == df2_to_check.values
    rows, cols = np.where(~comparison_values)
    return rows, cols, df1_to_check, df2_to_check


def preparation_output_file_whole(wb_output, sheets_without_pair, value_or_formula):
    status_sheet = rename_sheet(wb_output)
    status_sheet.cell(row=1,
                      column=1).value = "Sheets are compared by {}".format(value_or_formula)
    write_sheets_wo_pairs(status_sheet, sheets_without_pair)


def preparation_output_file_selected(wb_output, wb1_name, wb2_name, selected_sheets, value_or_formula):
    status_sheet = rename_sheet(wb_output)
    status_sheet.cell(row=1,
                      column=1).value = "Sheets are compared by {}".format(value_or_formula)
    status_sheet.cell(row=2, column=2).value = "Next sheets are compared:"
    status_sheet.cell(row=2, column=2).font = op.styles.Font(bold=True)
    status_sheet.cell(row=3, column=2).value = wb1_name
    status_sheet.cell(row=3, column=2).font = op.styles.Font(bold=True)
    status_sheet.cell(row=3, column=3).value = wb2_name
    status_sheet.cell(row=3, column=3).font = op.styles.Font(bold=True)
    for i, sheets in enumerate(selected_sheets):
        status_sheet.cell(row=4+i, column=2).value = sheets[0]
        status_sheet.cell(row=4+i, column=3).value = sheets[1]


def rename_sheet(wb_output):
    status_sheet_name = "Run"
    wb_output["Sheet"].title = status_sheet_name
    return wb_output[status_sheet_name]


def set_value_or_formula(wb1_path, wb2_path, value_or_formula):
    if value_or_formula == "value":
        wb1, wb2, wb_output = open_excel_files_values(wb1_path, wb2_path)
        wb1_v = None
        wb2_v = None
        formula = False
    elif value_or_formula == "formula":
        wb1, wb2, wb1_v, wb2_v, wb_output = open_excel_files_formulas(wb1_path, wb2_path)
        formula = True
    else:
        raise ValueError
    return wb1, wb2, wb1_v, wb2_v, wb_output, formula


def check_shape_and_set_to_one_shape(df1_to_check, df2_to_check, max_row, max_column):
    shape1 = df1_to_check.shape
    shape2 = df2_to_check.shape
    if shape1 != shape2:
        if shape1[0] < shape2[0]:
            insert_ = pd.DataFrame([[""] * shape1[1]], columns=df1_to_check.columns)
            diff = shape2[0] - shape1[0]
            for i in range(diff):
                df1_to_check = pd.concat([df1_to_check, insert_], ignore_index=True)
        else:
            insert_ = pd.DataFrame([[""] * shape2[1]], columns=df2_to_check.columns)
            diff = shape1[0] - shape2[0]
            for i in range(diff):
                df2_to_check = pd.concat([df2_to_check, insert_], ignore_index=True)

        if shape1[1] < shape2[1]:
            diff = shape2[1] - shape1[1]
            for i in range(diff):
                df1_to_check[shape1[1] + i] = ""
        else:
            diff = shape1[1] - shape2[1]
            for i in range(diff):
                df2_to_check[shape2[1] + i] = ""
    df1_to_check = df1_to_check.head(max_row)
    df2_to_check = df2_to_check.head(max_row)
    df1_to_check = df1_to_check.iloc[:, :max_column]
    df2_to_check = df2_to_check.iloc[:, :max_column]
    return df1_to_check, df2_to_check


class FileBroken(Exception):
    pass


def open_excel_files_values(wb1_path, wb2_path):
    try:
        wb1 = op.load_workbook(filename=wb1_path, read_only=True, data_only=True)
    except TypeError:
        raise FileBroken(f"File {wb1_path} is broken, resave it and rerun script")
    try:
        wb2 = op.load_workbook(filename=wb2_path, read_only=True, data_only=True)
    except TypeError:
        raise FileBroken(f"File {wb2_path} is broken, resave it and rerun script")

    wb_output = op.Workbook()
    return wb1, wb2, wb_output


def open_excel_files_formulas(wb1_path, wb2_path):
    wb1_v, wb2_v, wb_output = open_excel_files_values(wb1_path, wb2_path)
    try:
        wb1_f = op.load_workbook(filename=wb1_path, read_only=True, data_only=False)
    except TypeError:
        raise FileBroken(f"File {wb1_path} is broken, resave it and rerun script")
    try:
        wb2_f = op.load_workbook(filename=wb2_path, read_only=True, data_only=False)
    except TypeError:
        raise FileBroken(f"File {wb2_path} is broken, resave it and rerun script")
    return wb1_f, wb2_f, wb1_v, wb2_v, wb_output


def find_sheet_to_compare(wb1, wb2, wb1_name, wb2_name):
    ws1 = set(wb1.sheetnames)
    ws2 = set(wb2.sheetnames)
    sheets_to_check = list(ws1.intersection(ws2))
    output_list = []
    for sheet in wb1.sheetnames:
        if sheet in sheets_to_check:
            output_list.append(sheet)
    sheets_without_pair = {wb1_name: list(ws1-ws2), wb2_name: list(ws2-ws1)}
    return output_list, sheets_without_pair


def write_sheets_wo_pairs(status_sheet, dict_w_sheets):
    status_sheet.cell(row=2, column=2).value = "Sheets found in files that without pair (i.e. sheets that out of checking):"
    status_sheet.cell(row=2, column=2).font = op.styles.Font(bold=True)
    # status_sheet.merge_cells('B2:C2')
    for i, wb in enumerate(dict_w_sheets.keys()):
        status_sheet.cell(row=3, column=2+i).value = wb
        status_sheet.cell(row=3, column=2+i).font = op.styles.Font(bold=True)
        if not dict_w_sheets[wb]:
            status_sheet.cell(row=4, column=2+i).value = "All sheets are with pairs"
        else:
            for j, sh in enumerate(dict_w_sheets[wb]):
                status_sheet.cell(row=4+j, column=2+i).value = sh


def find_all_sheets(wb1_path, wb2_path):
    wb1_name = os.path.basename(wb1_path)
    wb2_name = os.path.basename(wb2_path)
    wb1 = op.load_workbook(filename=wb1_path, read_only=True, data_only=True)
    wb2 = op.load_workbook(filename=wb2_path, read_only=True, data_only=True)
    output_list = {wb1_name: wb1.sheetnames, wb2_name: wb2.sheetnames}
    return output_list, wb1_name, wb2_name


def setup_gui():
    # Creation of GUI
    def run_program_whole(e1, e2, e3, e4, e5, e6):
        run_flag = True
        if not all((e1.get(), e2.get(), e3.get(), e4.get(), e5.get(), e6.get())):
            messagebox.showerror("Error!", "Please provide all inputs!")
            run_flag = False
        if e1.get() == e2.get():
            messagebox.showerror("Error!", "You enter two identical files")
            run_flag = False
        if run_flag:
            root.withdraw()
            start_time = time.time()
            try:
                find_difference_whole(e1.get(), e2.get(), e3.get(), e4.get(), int(e5.get()), int(e6.get()))
                print("--- %s seconds ---" % (time.time() - start_time))
                messagebox.showinfo("Files are compared", "Files are compared successfully.")
                exit_()
                m.setup_gui()
            except FileBroken as FB:
                messagebox.showerror("Error!", str(FB))
                root.quit()
                compare_excels()

    def run_program_selected(e1, e2, e3, e4, e5, e6, e7):
        run_flag = True
        if not all((e1.get(), e2.get(), e3.get(), e5.get(), e6.get(), e7.get())):
            messagebox.showerror("Error!", "Please provide all inputs!")
            run_flag = False
        if not e4:
            messagebox.showerror("Error!", "Choose at least one sheet from each Excel")
            run_flag = False
        if run_flag:
            root.withdraw()
            start_time = time.time()
            try:
                find_difference_selected(e1.get(), e2.get(), e3.get(), e4, e5.get(), int(e6.get()), int(e7.get()))
                print("--- %s seconds ---" % (time.time() - start_time))
                messagebox.showinfo("Files are compared", "Files are compared successfully.")
                exit_()
                m.setup_gui()
            except FileBroken as FB:
                messagebox.showerror("Error!", str(FB))
                root.quit()
                compare_sheets()

    def exit_():
        root.destroy()
        root.quit()

    def get_excel1_path(e1):
        path = filedialog.askopenfilename(title="Select first Excel file",
                                          filetypes=(("First Excel", "*.xls *.xlsx *.xlsm"), ("all files", "*.*")))
        e1.delete(0, tk.END)
        e1.insert(0, path)

    def get_excel2_path(e2):
        path = filedialog.askopenfilename(title="Select second Excel file",
                                          filetypes=(("Second Excel", "*.xls *.xlsx *.xlsm"), ("all files", "*.*")))
        e2.delete(0, tk.END)
        e2.insert(0, path)

    def get_output_path(e3):
        path = filedialog.askdirectory(title="Select directory for output")
        e3.delete(0, tk.END)
        e3.insert(0, path)

    root = tk.Tk()
    root.title("Find difference")
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    window_width = 400
    window_height = 100
    center_x = int(screen_width / 2 - window_width / 2)
    center_y = int(screen_height / 2 - window_height / 2)
    root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

    def compare_excels():
        root.withdraw()
        newWindow = tk.Toplevel(root)
        tk.Label(newWindow,
                 text="Compare sheets with the same name").grid(row=0, columnspan=2, sticky=tk.S)
        tk.Label(newWindow,
                 text="First Excel:").grid(row=1, columnspan=2, sticky=tk.S)
        tk.Label(newWindow,
                 text="Second Excel:").grid(row=3, columnspan=2)
        tk.Label(newWindow,
                 text="Output location:").grid(row=5, columnspan=2)
        tk.Label(newWindow,
                 text="What to check:").grid(row=7, columnspan=2)
        tk.Label(newWindow,
                 text="Input number of rows for checking:").grid(row=9, columnspan=2)
        tk.Label(newWindow,
                 text="Input number of columns for checking:").grid(row=11, columnspan=2)

        e1 = tk.Entry(newWindow, exportselection=False)
        e2 = tk.Entry(newWindow, exportselection=False)
        e3 = tk.Entry(newWindow, exportselection=False)
        e3.insert(tk.END, r"C:/Temp")
        e4 = tk.StringVar(newWindow)
        e4.set("value")
        e5 = tk.Entry(newWindow, exportselection=False)
        e5.insert(tk.END, r"2000")
        e6 = tk.Entry(newWindow, exportselection=False)
        e6.insert(tk.END, r"200")
        e1.focus_set()

        e1.grid(row=2, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        e2.grid(row=4, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        e3.grid(row=6, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        tk.OptionMenu(newWindow, e4, "value", "formula").grid(row=8, columnspan=2, sticky=tk.E + tk.W)
        e5.grid(row=10, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        e6.grid(row=12, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        tk.Button(newWindow, text='...', width=4,
                  command=lambda: get_excel1_path(e1)).grid(row=2, column=1, sticky=tk.W)
        tk.Button(newWindow, text='...', width=4,
                  command=lambda: get_excel2_path(e2)).grid(row=4, column=1, sticky=tk.W)
        tk.Button(newWindow, text='...', width=4,
                  command=lambda: get_output_path(e3)).grid(row=6, column=1, sticky=tk.W)

        hbox = tk.Frame(newWindow)
        tk.Button(hbox, text='Run', width=12,
                  command=lambda: run_program_whole(e1, e2, e3, e4, e5, e6)).pack(side=tk.LEFT, expand=True)
        tk.Button(hbox, text='Exit', width=12,
                  command=exit_).pack(side=tk.LEFT, expand=True)
        hbox.grid(row=13, column=0, columnspan=2, sticky=tk.E + tk.W)

        newWindow.rowconfigure(0, pad=15)
        newWindow.rowconfigure(13, pad=20)
        newWindow.columnconfigure(0, weight=1, minsize=600)
        newWindow.columnconfigure(1, pad=5)
        newWindow.geometry('+%d+%d' % (newWindow.winfo_screenwidth()/2 - 150, newWindow.winfo_screenheight()/2 - 300))
        newWindow.update()
        newWindow.minsize(newWindow.winfo_width(), newWindow.winfo_height())

    def compare_sheets():
        root.withdraw()
        newWindow = tk.Toplevel(root)
        tk.Label(newWindow,
                 text="Compare selected sheets:").grid(row=0, columnspan=2, sticky=tk.S)
        tk.Label(newWindow,
                 text="First Excel:").grid(row=1, columnspan=2, sticky=tk.S)
        tk.Label(newWindow,
                 text="Second Excel:").grid(row=3, columnspan=2)
        tk.Label(newWindow,
                 text="Output location:").grid(row=5, columnspan=2)
        tk.Label(newWindow,
                 text="What to check:").grid(row=7, columnspan=2)
        tk.Label(newWindow,
                 text="Input number of rows for checking:").grid(row=9, columnspan=2)
        tk.Label(newWindow,
                 text="Input number of columns for checking:").grid(row=11, columnspan=2)
        tk.Label(newWindow,
                 text="Select sheets:").grid(row=14, columnspan=2)

        e1 = tk.Entry(newWindow, exportselection=False)
        e2 = tk.Entry(newWindow, exportselection=False)
        e3 = tk.Entry(newWindow, exportselection=False)
        e3.insert(tk.END, r"C:/Temp")
        e5 = tk.StringVar(newWindow)
        e5.set("value")
        e6 = tk.Entry(newWindow, exportselection=False)
        e6.insert(tk.END, r"2000")
        e7 = tk.Entry(newWindow, exportselection=False)
        e7.insert(tk.END, r"200")
        e1.focus_set()

        e1.grid(row=2, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        e2.grid(row=4, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        e3.grid(row=6, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        tk.OptionMenu(newWindow, e5, "value", "formula").grid(row=8, columnspan=2, sticky=tk.E + tk.W)
        e6.grid(row=10, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        e7.grid(row=12, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
        tk.Button(newWindow, text='...', width=4,
                  command=lambda: get_excel1_path(e1)).grid(row=2, column=1, sticky=tk.W)
        tk.Button(newWindow, text='...', width=4,
                  command=lambda: get_excel2_path(e2)).grid(row=4, column=1, sticky=tk.W)
        tk.Button(newWindow, text='...', width=4,
                  command=lambda: get_output_path(e3)).grid(row=6, column=1, sticky=tk.W)

        hbox = tk.Frame(newWindow)
        tk.Button(hbox, text='Continue', width=12,
                  command=lambda: choose_sheets(e1, e2)).pack(side=tk.LEFT, expand=True)
        tk.Button(hbox, text='Exit', width=12,
                  command=exit_).pack(side=tk.LEFT, expand=True)
        hbox.grid(row=15, column=0, columnspan=2, sticky=tk.E + tk.W)

        newWindow.rowconfigure(0, pad=15)
        newWindow.rowconfigure(15, pad=20)
        newWindow.columnconfigure(0, weight=1, minsize=600)
        newWindow.columnconfigure(1, pad=5)

        newWindow.geometry('+%d+%d' % (newWindow.winfo_screenwidth()/2 - 150,
                                       newWindow.winfo_screenheight()/2 - 300))
        newWindow.update()
        newWindow.minsize(newWindow.winfo_width(), newWindow.winfo_height())

        def choose_sheets(e1, e2):
            newWindow.withdraw()
            e4 = []

            def add_item():
                sh1 = wb1.curselection()
                sh2 = wb2.curselection()
                e4.append([wb1.get(sh1), wb2.get(sh2)])
                output.insert(tk.END, "Sheet \"{}\" will be compared with sheet \"{}\"".format(wb1.get(sh1), wb2.get(sh2)))
                wb1.selection_clear(0, tk.END)
                wb2.selection_clear(0, tk.END)

            sheets, wb1_name, wb2_name = find_all_sheets(e1.get(), e2.get())

            sheet_windows = tk.Toplevel(newWindow)
            tk.Label(sheet_windows, text="Select sheets to check").grid(row=0, columnspan=2, sticky=tk.S)
            tk.Label(sheet_windows, text=wb1_name).grid(row=1, column=0, sticky=tk.S)
            tk.Label(sheet_windows, text=wb2_name).grid(row=1, column=1, sticky=tk.S)
            tk.Label(sheet_windows, text="Next sheets will be compared:").grid(row=3, columnspan=2)

            wb1 = tk.Listbox(sheet_windows, exportselection=0)
            for sheet in sheets[wb1_name]:
                wb1.insert(tk.END, sheet)
            wb2 = tk.Listbox(sheet_windows, exportselection=0)
            for sheet in sheets[wb2_name]:
                wb2.insert(tk.END, sheet)
            output = tk.Listbox(sheet_windows)

            wb1.grid(row=2, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
            wb2.grid(row=2, column=1, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
            output.grid(row=4, columnspan=2, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)

            hbox_nw = tk.Frame(sheet_windows)
            tk.Button(hbox_nw, text='Select', width=12,
                      command=add_item).pack(side=tk.LEFT, expand=True)
            tk.Button(hbox_nw, text='Run', width=12,
                      command=lambda: run_program_selected(e1, e2, e3, e4, e5, e6, e7)).pack(side=tk.LEFT, expand=True)
            tk.Button(hbox_nw, text='Exit', width=12,
                      command=exit_).pack(side=tk.LEFT, expand=True)
            hbox_nw.grid(row=5, column=0, columnspan=2, sticky=tk.E + tk.W)

            sheet_windows.rowconfigure(0, pad=15)
            sheet_windows.rowconfigure(5, pad=20)
            sheet_windows.columnconfigure(0, weight=1, minsize=200)
            sheet_windows.columnconfigure(1, pad=5)
            sheet_windows.geometry('+%d+%d' % (sheet_windows.winfo_screenwidth() / 2 - 100,
                                               sheet_windows.winfo_screenheight() / 2 - 300))
            sheet_windows.update()
            sheet_windows.minsize(sheet_windows.winfo_width(), sheet_windows.winfo_height())

    # tk.Button(root, text='Find differences for sheets with the same names from two files', command=compare_excels).pack(pady=50)
    # tk.Button(root, text='Find differences for selected sheets from two files', command=compare_sheets).pack(pady=150)
    btn1 = tk.Button(
        root,
        text='Find differences for sheets with the same names from two files',
        command=compare_excels,
        padx=15,
        pady=5
    )

    btn1.pack(expand=True, side=tk.TOP)

    btn2 = tk.Button(
        root,
        text='Find differences for selected sheets from two files',
        command=compare_sheets,
        padx=20,
        pady=5
    )
    btn2.pack(expand=True, side=tk.BOTTOM)

    tk.mainloop()


def get_arguments():
    parser = argparse.ArgumentParser(description="WP135 Stress Tool")
    parser.add_argument("-i", required=False, metavar='FIRST_EXCEL',
                        help="path to the first excel file, including file name")
    parser.add_argument("-e", required=False, metavar='SECOND_EXCEL',
                        help="path to the second file, including file name")
    parser.add_argument("-o", required=False, metavar='OUTPUT_PATH',
                        help="where the output directories and files will be created")
    parser.add_argument("-k", required=False, metavar='VALUE_FORMULA',
                        help="What to check: by values or by formulas?")
    parser.add_argument("-j", required=False, metavar='MAX_ROWS',
                        help="Max rows to check")
    parser.add_argument("-p", required=False, metavar='MAX_COLUMNS',
                        help="Max columns to check?")
    return parser.parse_args()


def start(root):
    root.destroy()
    args = get_arguments()
    if all([args.i, args.e, args.o, args.k, args.j, args.p]):
        find_difference_whole(args.i, args.e, args.o, args.k, args.j, args.p)
    else:
        setup_gui()
        