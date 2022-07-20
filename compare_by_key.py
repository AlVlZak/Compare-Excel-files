# This is a script to compare two definition Excel files.
__version__ = "1.0"
import argparse
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import os
from os.path import exists
import logging
import time
from pathlib import Path

import pandas as pd
import numpy as np
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows

import compare_excels as fd
import main as m

WARNING_SHEET = dict()


def compare_definition(wb1_path, wb2_path, output_path, keys_path, value_or_formula="value"):
    wb1_name = os.path.basename(wb1_path)
    wb2_name = os.path.basename(wb2_path)
    wb_output_name = "Comp_by_key_results.xlsx"
    separator = "--->"
    logging.basicConfig(filename=os.path.join(output_path, "logs.log"),
                        encoding='utf-8',
                        level=logging.DEBUG,
                        format='\n%(levelname)s: %(asctime)s: %(message)s',
                        filemode="w",
                        datefmt='%Y/%m/%d %H:%M:%S',
                        )
    wb1, wb2, wb1_v, wb2_v, wb_output, formula = set_value_or_formula(wb1_path, wb2_path, value_or_formula)
    keys_loc = pd.read_csv(keys_path)
    list_of_keys = keys_loc['sheet_name'].tolist()
    sheets_to_check, sheets_without_pair, output_list_not_def = find_sheet_to_compare(wb1, wb2, wb1_name, wb2_name, list_of_keys)
    preparation_output_file(wb_output, sheets_without_pair, value_or_formula)
    keys_loc = keys_loc.set_index("sheet_name")
    len_of_sheets_to_check = len(sheets_to_check)
    for k, sheet in enumerate(sheets_to_check, 1):
        max_row = max(wb1[sheet].max_row, wb2[sheet].max_row)
        max_column = max(wb1[sheet].max_column, wb2[sheet].max_column)
        row_id = keys_loc.row_ID.loc[sheet] - 1
        column_id = keys_loc.column_ID.loc[sheet] - 1
        df1_to_check, df2_to_check, wb_res, sheet = define_sheet(wb1, wb2, wb_output, sheet, formula, column_id, row_id,
                                                                 wb1_v, wb2_v)
        if "CommonStyle" in str(df1_to_check.iloc[0, 0]):
            to_print = "Sheet \"{}\" is in Common style and was compared cell by cell".format(sheet)
            print(to_print)
            logging.warning(to_print)
            finddiff.find_difference(wb1, wb2, wb_output, sheet, max_row, max_column, len_of_sheets_to_check, k, separator, formula, wb1_v, wb2_v)
        else:
            print_status(sheet, k, max_row, max_column, len_of_sheets_to_check)
            comparison_values, keys, unique_keys_df1, unique_keys_df2, df1_to_check, df2_to_check, header, df_unique = \
                compare_sheet(df1_to_check, df2_to_check, column_id, row_id, sheet)
            if formula:
                df1_to_check, df2_to_check = replace_formulas_by_values(wb1_v[sheet], wb2_v[sheet], column_id)
            rows = filling_output_df(df1_to_check, df2_to_check, separator, comparison_values, keys, unique_keys_df1, unique_keys_df2, header, wb1_name, wb2_name)
            add_comparing_values_to_output(wb_res, rows, separator, header)
    len_of_sheets_to_check = len(output_list_not_def)
    to_print = "Comparing cell by cell sheets, that are not in Common style and not in list of column ID table"
    print(to_print)
    logging.info(to_print)
    for k, sheet in enumerate(output_list_not_def, 1):
        max_row = max(wb1[sheet].max_row, wb2[sheet].max_row)
        max_column = max(wb1[sheet].max_column, wb2[sheet].max_column)
        df1_to_check, df2_to_check, wb_res, sheet = define_sheet(wb1, wb2, wb_output, sheet, formula, "", "", wb1_v, wb2_v)
        fd.find_difference(wb1, wb2, wb_output, sheet, max_row, max_column, len_of_sheets_to_check, k, separator,
                                 formula, wb1_v, wb2_v)
    list_of_keys.extend(output_list_not_def)
    wb_output = color_and_add_hyperlinks(wb_output, list_of_keys, wb1, wb2)
    add_warnings_to_output(wb_output)
    print("\nCreating output file")
    logging.info("Creating output file")
    output_name = os.path.join(output_path, wb_output_name)
    wb_output.save(output_name)
    if formula:
        wb1_v.close()
        wb2_v.close()
    wb1.close()
    wb2.close()


def find_sheet_to_compare(wb1, wb2, wb1_name, wb2_name, keys):
    ws1 = set(wb1.sheetnames)
    ws2 = set(wb2.sheetnames)
    sheets_to_check = list(ws1.intersection(ws2))
    output_list = []
    for sheet in wb1.sheetnames:
        if sheet in sheets_to_check and sheet in keys:
            output_list.append(sheet)
    sh_only_in_first_wb = ws1-ws2
    sh_only_in_second_wb = ws2-ws1
    sh_only_in_first_wb.difference_update(keys)
    sh_only_in_second_wb.difference_update(keys)
    sheets_without_pair = {wb1_name: list(sh_only_in_first_wb), wb2_name: list(sh_only_in_second_wb)}
    output_list_not_def = []
    for sheet in wb1.sheetnames:
        if sheet in sheets_to_check and sheet not in output_list:
            output_list_not_def.append(sheet)
    return output_list, sheets_without_pair, output_list_not_def


def set_value_or_formula(wb1_path, wb2_path, value_or_formula):
    if value_or_formula == "value":
        wb1, wb2, wb_output = open_excel_files_values(wb1_path, wb2_path)
        wb1_v = None
        wb2_v = None
        formula = False
    elif value_or_formula == "formula":
        wb1, wb2, wb1_v, wb2_v, wb_output = open_excel_files_formulas(wb1_path, wb2_path)
        formula = True
    else:
        raise ValueError
    return wb1, wb2, wb1_v, wb2_v, wb_output, formula


class FileBroken(Exception):
    pass


def open_excel_files_values(wb1_path, wb2_path):
    try:
        wb1 = op.load_workbook(filename=wb1_path, read_only=True, data_only=True)
    except TypeError:
        raise FileBroken(f"File {wb1_path} is broken, resave it and rerun script")
    try:
        wb2 = op.load_workbook(filename=wb2_path, read_only=True, data_only=True)
    except TypeError:
        raise FileBroken(f"File {wb2_path} is broken, resave it and rerun script")

    wb_output = op.Workbook()
    return wb1, wb2, wb_output


def open_excel_files_formulas(wb1_path, wb2_path):
    wb1_v, wb2_v, wb_output = open_excel_files_values(wb1_path, wb2_path)
    try:
        wb1_f = op.load_workbook(filename=wb1_path, read_only=True, data_only=False)
    except TypeError:
        raise FileBroken(f"File {wb1_path} is broken, resave it and rerun script")
    try:
        wb2_f = op.load_workbook(filename=wb2_path, read_only=True, data_only=False)
    except TypeError:
        raise FileBroken(f"File {wb2_path} is broken, resave it and rerun script")
    return wb1_f, wb2_f, wb1_v, wb2_v, wb_output


def preparation_output_file(wb_output, sheets_without_pair, value_or_formula):
    status_sheet = rename_sheet(wb_output)
    status_sheet.cell(row=1,
                      column=1).value = "Sheets are compared by {}".format(value_or_formula)
    write_sheets_wo_pairs(status_sheet, sheets_without_pair)


def write_sheets_wo_pairs(status_sheet, dict_w_sheets):
    status_sheet.cell(row=2, column=2).value = "Sheets found in files that without pair (i.e. sheets that out of checking):"
    status_sheet.cell(row=2, column=2).font = op.styles.Font(bold=True)
    status_sheet.merge_cells('B2:C2')
    status_sheet.cell(row=2, column=2).alignment = op.styles.Alignment(horizontal='center')
    for i, wb in enumerate(dict_w_sheets.keys()):
        status_sheet.cell(row=3, column=2+i).value = wb
        status_sheet.cell(row=3, column=2+i).font = op.styles.Font(bold=True)
        if not dict_w_sheets[wb]:
            status_sheet.cell(row=4, column=2+i).value = "All sheets are with pairs"
        else:
            for j, sh in enumerate(dict_w_sheets[wb]):
                status_sheet.cell(row=4+j, column=2+i).value = sh


def rename_sheet(wb_output):
    status_sheet_name = "Run"
    wb_output["Sheet"].title = status_sheet_name
    return wb_output[status_sheet_name]


def define_sheet(wb1, wb2, wb_output, sheet, formula, column_id, row_id, wb1_v, wb2_v):
    wb_output.create_sheet(sheet)
    wb1_to_check = wb1[sheet]
    wb2_to_check = wb2[sheet]
    wb_res = wb_output[sheet]
    df1_to_check = pd.DataFrame(wb1_to_check.values)
    df2_to_check = pd.DataFrame(wb2_to_check.values)
    if row_id != "":
        if (row_id < df1_to_check.shape[0] and column_id < df1_to_check.shape[1]) and df1_to_check.iloc[row_id, column_id] is None:
            df1_to_check = df1_to_check.iloc[:row_id]
        if (row_id < df2_to_check.shape[0] and column_id < df2_to_check.shape[1]) and df2_to_check.iloc[row_id, column_id] is None:
            df2_to_check = df2_to_check.iloc[:row_id]
    if formula and column_id != "":
        df1_temp = pd.DataFrame(wb1_v[sheet].values)
        df2_temp = pd.DataFrame(wb2_v[sheet].values)
        df1_to_check[column_id] = df1_temp[column_id]
        df2_to_check[column_id] = df2_temp[column_id]
        del df1_temp, df2_temp
    return df1_to_check, df2_to_check, wb_res, sheet


def compare_sheet(df1_to_check, df2_to_check, column_id, row_id, sheet):
    header = df1_to_check.iloc[row_id - 1].loc[column_id]
    df1_to_check = df1_to_check.set_index(column_id)
    df2_to_check = df2_to_check.set_index(column_id)
    df1_keys = df1_to_check.iloc[row_id-1:].index
    df2_keys = df2_to_check.iloc[row_id-1:].index
    global WARNING_SHEET
    WARNING_SHEET[sheet] = {"File1": "", "File2": ""}
    if not df1_to_check.index.is_unique:
        dup_value = df1_to_check.index[df1_to_check.index.duplicated(keep='first')].unique().to_list()
        dup_value = [x for x in dup_value if x not in [None, np.nan]]
        dup_value = ", ".join(dup_value)
        text = f"There are duplicate values in index in the first file on sheet \"{sheet}\", first occurrence will be used\n" \
               f"Duplicate indexes in first file are: {dup_value}"
        if dup_value:
            WARNING_SHEET[sheet]["File1"] = dup_value
            print(text)
            logging.warning(text)
        df1_to_check = df1_to_check[~df1_to_check.index.duplicated(keep='first')]
    if not df2_to_check.index.is_unique:
        dup_value = df2_to_check.index[df2_to_check.index.duplicated(keep='first')].unique().to_list()
        dup_value = [x for x in dup_value if x not in [None, np.nan]]
        dup_value = ", ".join(dup_value)
        text = f"There are duplicate values in index in the second file on sheet \"{sheet}\", first occurrence will be used\n" \
               f"Duplicate indexes in second file are: {dup_value}"
        if dup_value:
            WARNING_SHEET[sheet]["File2"] = dup_value
            print(text)
            logging.warning(text)
        df2_to_check = df2_to_check[~df2_to_check.index.duplicated(keep='first')]
    if WARNING_SHEET[sheet]["File1"] == "" and WARNING_SHEET[sheet]["File2"] == "":
        WARNING_SHEET.pop(sheet, None)
    if isinstance(df1_keys, pd.MultiIndex):
        unique_keys_df1 = df1_keys.difference(df2_keys)
        unique_keys_df2 = df2_keys.difference(df1_keys)
    else: # (df1_keys.size >= 1 and df1_keys[1] != "None") or (df2_keys.size >= 1 and df2_keys[1] != "None"):
        unique_keys_df1 = np.setdiff1d(df1_keys, df2_keys)
        unique_keys_df2 = np.setdiff1d(df2_keys, df1_keys)
    comparison_values = df1_to_check.copy()
    comparison_values = pd.concat([comparison_values, df2_to_check[df2_to_check.index.isin(unique_keys_df2)]])
    keys = (df1_keys.intersection(df2_keys)).tolist()
    for key in keys:
        comparison_values.loc[key] = df1_to_check.loc[key].values == df2_to_check.loc[key].values
    df_unique = comparison_values[comparison_values.index.isin(unique_keys_df1) | comparison_values.index.isin(unique_keys_df2)]
    comparison_values = comparison_values[~comparison_values.index.isin(unique_keys_df1) & ~comparison_values.index.isin(unique_keys_df2)]
    comparison_values = comparison_values.astype(bool)
    return comparison_values, keys, unique_keys_df1, unique_keys_df2, df1_to_check, df2_to_check, header, df_unique


def print_status(sheet, k, max_row, max_column, len_of_sheets_to_check):
    status = f"Working with sheet \"{sheet}\", max row is {max_row}, max column is {max_column},  " \
             f"progress {k}/{len_of_sheets_to_check}"
    print(status)
    logging.info(status)


def replace_formulas_by_values(wb1_v, wb2_v, column_id):
    df1_to_check = pd.DataFrame(wb1_v.values)
    df2_to_check = pd.DataFrame(wb2_v.values)
    df1_to_check = df1_to_check.set_index(column_id)
    df2_to_check = df2_to_check.set_index(column_id)
    if not df1_to_check.index.is_unique:
        df1_to_check = df1_to_check[~df1_to_check.index.duplicated(keep='first')]
    if not df2_to_check.index.is_unique:
        df2_to_check = df2_to_check[~df2_to_check.index.duplicated(keep='first')]
    return df1_to_check, df2_to_check


def filling_output_df(df1_to_check, df2_to_check, separator, comparison_values, keys, unique_keys_df1, unique_keys_df2, header, wb1_name, wb2_name):
    output_df = comparison_values.copy()
    status_column = max(list(comparison_values.columns)) + 2
    output_df[status_column] = ""
    columns = list(comparison_values.columns)
    empty_cell = [f"nan {separator} nan", f"None {separator} ", f"nan {separator} ", f" {separator} None",
                  f" {separator} nan"]
    for key in keys:
        if key == header:
            output_df.loc[key] = df1_to_check.loc[key]
            output_df.at[key, status_column] = "STATUS"
            continue
        check_list = comparison_values.loc[key].tolist()
        df1_list = df1_to_check.loc[key].values.tolist()
        df2_list = df2_to_check.loc[key].values.tolist()
        status = True
        for i, elm in enumerate(check_list):
            if elm:
                output = '{}'.format(df1_list[i])
            else:
                output = '{} {} {}'.format(df1_list[i], separator, df2_list[i])
                status = False
            if output in empty_cell:
                output = ""
            output_df.at[key, columns[i]] = output
        if status:
            output_df.at[key, status_column] = "SIMILAR"
        else:
            output_df.at[key, status_column] = "DIFFERENT"
    output_df, status_column = fill_df_with_unique_data(output_df, df1_to_check, unique_keys_df1, status_column, wb1_name)
    output_df, status_column = fill_df_with_unique_data(output_df, df2_to_check, unique_keys_df2, status_column, wb2_name)
    output_df.replace(to_replace="None", value="", inplace=True)  # , regex=True)
    output_df.replace(to_replace=True, value="", inplace=True)
    output_df.replace(to_replace=False, value="", inplace=True)
    output_df.reset_index(inplace=True)
    output_df = output_df.sort_index(ascending=True, axis=1)
    rows = dataframe_to_rows(output_df, index=False, header=False)
    return rows


def fill_df_with_unique_data(df_output, df_data, unique_keys, status_column, file_name):
    for key in unique_keys:
        df_output.loc[key] = df_data.loc[key]
        df_output.at[key, status_column] = f"ONLY_FOUND_IN_{file_name}"
    return df_output, status_column


def add_comparing_values_to_output(wb_res, rows, separator, header):
    color_red = False
    color_blue = False
    value_starts = False
    header_starts = False
    color_dict = {"green": "99ffcc", "red": "ff9999", "blue": "33ccff"}
    for r_idx, row in enumerate(rows, 1):
        color_for_not_found = False
        color = color_dict["green"]
        for c_idx, value in enumerate(row, 1):
            if value != "None":
                wb_res.cell(row=r_idx, column=c_idx, value=value)
            if str(value) == header:
                header_starts = True
            if header_starts and str(value) == "STATUS":
                value_starts = True
                header_starts = False
                break
            if value_starts:
                if "ONLY_FOUND_IN_" in row[-1] and not color_for_not_found:
                    color = color_dict["blue"]
                    color_for_not_found = True
                    color_blue = True
                elif not color_for_not_found:
                    if separator in str(value) or str(value) == "DIFFERENT":
                        color = color_dict["red"]
                        color_red = True
                    else:
                        color = color_dict["green"]
                wb_res.cell(row=r_idx, column=c_idx).fill = op.styles.PatternFill("solid", start_color=color)
    if color_red:
        wb_res.sheet_properties.tabColor = color_dict["red"]
    elif color_blue:
        wb_res.sheet_properties.tabColor = color_dict["blue"]
    else:
        wb_res.sheet_properties.tabColor = color_dict["green"]

    if wb_res.cell(row=1, column=1).value == "" or wb_res.cell(row=1, column=1).value is None:
        wb_res.cell(row=1, column=1, value="Go to sheet Run")
    wb_res.cell(row=1, column=1).hyperlink = "#Run!A1"
    wb_res.cell(row=1, column=1).style = "Hyperlink"

    to_break = False
    for row in wb_res.iter_rows():
        if to_break:
            break
        for cell in row:
            if cell.value == "STATUS":
                wb_res.auto_filter.ref = f"A{cell.row}:{cell.column_letter}{wb_res.max_row}"
                wb_res.auto_filter.add_filter_column(cell.column - 1, ["ONLY_FOUND_IN_*", "DIFFERENT"])
                to_break = True
                break


def create_hyperlink(ws, row_id, column_id, link):
    ws.cell(row=row_id, column=column_id).hyperlink = f"#\'{link}\'!A1"
    ws.cell(row=row_id, column=column_id, value=f"Go to sheet \"{link}\"")
    ws.cell(row=row_id, column=column_id).style = "Hyperlink"


def color_and_add_hyperlinks(wb, sheets, wb1, wb2):
    run_sheet = wb["Run"]
    col = 5
    run_sheet.cell(row=2, column=col).value = "Compared sheets"
    run_sheet.cell(row=2, column=col).font = op.styles.Font(bold=True)
    wb1_sheets = wb1.sheetnames
    wb2_sheets = wb2.sheetnames
    color_dict = {"grey": "C2C0BF", "dark_grey": "6C6B6B"}
    for i, sheet in enumerate(sheets, 3):
        if sheet in wb1_sheets and sheet in wb2_sheets:
            create_hyperlink(run_sheet, i, col, sheet)
            color = wb[sheet].sheet_properties.tabColor
        elif sheet in wb1_sheets:
            run_sheet.cell(row=i, column=col).value = f"Sheet \"{sheet}\" is only in first file"
            color = color_dict["grey"]
        elif sheet in wb2_sheets:
            run_sheet.cell(row=i, column=col).value = f"Sheet \"{sheet}\" is only in second file"
            color = color_dict["grey"]
        else:
            run_sheet.cell(row=i, column=col).value = f"Sheet \"{sheet}\" was not found in both files"
            color = color_dict["dark_grey"]
        run_sheet.cell(row=i, column=col).fill = op.styles.PatternFill("solid", start_color=color)
    autofit_column_width(run_sheet)
    return wb


def autofit_column_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def add_warnings_to_output(wb_output):
    run_sheet = wb_output["Run"]
    run_sheet.cell(row=1, column=6).value = f"Duplicate indexes"
    run_sheet.cell(row=1, column=6).font = op.styles.Font(bold=True)
    run_sheet.merge_cells('F1:G1')

    run_sheet.cell(row=2, column=6).value = run_sheet.cell(row=3, column=2).value
    run_sheet.cell(row=2, column=6).font = op.styles.Font(bold=True)
    run_sheet.cell(row=2, column=7).value = run_sheet.cell(row=3, column=3).value
    run_sheet.cell(row=2, column=7).font = op.styles.Font(bold=True)
    for row in run_sheet.iter_rows(min_row=3, min_col=5, max_col=5):
        for cell in row:
            sheet = cell.value.replace('Go to sheet "', "").replace('"', "")
            if sheet in WARNING_SHEET.keys():
                run_sheet.cell(row=cell.row, column=6).value = WARNING_SHEET[sheet]["File1"]
                run_sheet.cell(row=cell.row, column=7).value = WARNING_SHEET[sheet]["File2"]


def setup_gui():
    # Creation of GUI
    def run_program():
        run_flag = True
        if not all((e1.get(), e2.get(), e3.get(), e4.get(), e5.get())):
            messagebox.showerror("Error!", "Please provide all inputs!")
            run_flag = False
        if e1.get() == e2.get():
            messagebox.showerror("Error!", "You enter two identical files")
            run_flag = False
        if run_flag:
            root.withdraw()
            start_time = time.time()
            try:
                compare_definition(e1.get(), e2.get(), e3.get(), e4.get(), e5.get())
                output = "--- %s seconds ---" % (time.time() - start_time)
                print(output)
                logging.info(output)
                messagebox.showinfo("Files are compared", "Files are compared successfully. Check 'logs.log' for warnings")
                exit_()
                m.setup_gui()
            except FileBroken as FB:
                messagebox.showerror("Error!", str(FB))
                root.quit()
                setup_gui()
            except PermissionError:
                messagebox.showerror("Error!", 'File "Comp_def_results.xlsx" cannot be written to output folder.\n'
                                               'Maybe someone has the file open.\n'
                                               'Close the file and try again.')
                root.quit()
                setup_gui()

    def exit_():
        root.destroy()
        root.quit()

    def get_excel1_path():
        path = filedialog.askopenfilename(title="Select first definition file",
                                          filetypes=(("First definition file", "*.xls *.xlsx *.xlsm"), ("all files", "*.*")))
        e1.delete(0, tk.END)
        e1.insert(0, path)

    def get_excel2_path():
        path = filedialog.askopenfilename(title="Select second definition file",
                                          filetypes=(("Second definition file", "*.xls *.xlsx *.xlsm"), ("all files", "*.*")))
        e2.delete(0, tk.END)
        e2.insert(0, path)

    def get_output_path():
        path = filedialog.askdirectory(title="Select directory for output")
        e3.delete(0, tk.END)
        e3.insert(0, path)

    def get_keys_path():
        path = filedialog.askopenfilename(title="Select file with ID starting table",
                                          filetypes=(("ID starting table", "*.csv *.txt"), ("all files", "*.*")))
        e4.delete(0, tk.END)
        e4.insert(0, path)

    root = tk.Tk()
    root.title("Compare definition")

    tk.Label(root,
             text="Comparing two definition files").grid(row=0, columnspan=2, sticky=tk.S)
    tk.Label(root,
             text="First definition file:").grid(row=1, columnspan=2, sticky=tk.S)
    tk.Label(root,
             text="Second definition file:").grid(row=3, columnspan=2)
    tk.Label(root,
             text="Output location:").grid(row=5, columnspan=2)
    tk.Label(root,
             text="ID starting table:").grid(row=7, columnspan=2)

    tk.Label(root,
             text="What to check:").grid(row=9, columnspan=2)

    e1 = tk.Entry(root, exportselection=False)
    e2 = tk.Entry(root, exportselection=False)
    e3 = tk.Entry(root, exportselection=False)
    e3.insert(tk.END, r"C:/Temp")
    e4 = tk.Entry(root, exportselection=False)
    if exists(f"{Path(sys.argv[0]).parent.absolute()}\\keys.csv"):
        e4.insert(tk.END, f"{Path(sys.argv[0]).parent.absolute()}\\keys.csv")
    e5 = tk.StringVar(root)
    e5.set("value")

    e1.grid(row=2, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
    e2.grid(row=4, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
    e3.grid(row=6, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
    e4.grid(row=8, column=0, padx=5, pady=5, ipady=2, sticky=tk.E + tk.W)
    tk.OptionMenu(root, e5, "value", "formula").grid(row=10, columnspan=2, sticky=tk.E + tk.W)
    tk.Button(root, text='...', width=4,
              command=get_excel1_path).grid(row=2, column=1, sticky=tk.W)
    tk.Button(root, text='...', width=4,
              command=get_excel2_path).grid(row=4, column=1, sticky=tk.W)
    tk.Button(root, text='...', width=4,
              command=get_output_path).grid(row=6, column=1, sticky=tk.W)
    tk.Button(root, text='...', width=4,
              command=get_keys_path).grid(row=8, column=1, sticky=tk.W)

    hbox = tk.Frame(root)
    tk.Button(hbox, text='Run', width=12,
              command=run_program).pack(side=tk.LEFT, expand=True)
    tk.Button(hbox, text='Exit', width=12,
              command=exit_).pack(side=tk.LEFT, expand=True)
    hbox.grid(row=15, column=0, columnspan=2, sticky=tk.E + tk.W)

    root.rowconfigure(0, pad=15)
    root.rowconfigure(15, pad=20)
    root.columnconfigure(0, weight=1, minsize=600)
    root.columnconfigure(1, pad=5)
    root.geometry('+%d+%d' % (root.winfo_screenwidth()/2 - 150,  root.winfo_screenheight()/2 - 300))
    root.update()
    root.minsize(root.winfo_width(), root.winfo_height())
    tk.mainloop()


def get_arguments():
    parser = argparse.ArgumentParser(description="WP135 Stress Tool")
    parser.add_argument("-i", required=False, metavar='FIRST_FILE',
                        help="path to the first excel file, including file name")
    parser.add_argument("-e", required=False, metavar='SECOND_FILE',
                        help="path to the second file, including file name")
    parser.add_argument("-o", required=False, metavar='OUTPUT_PATH',
                        help="where the output directories and files will be created")
    parser.add_argument("-j", required=False, metavar='KEYS_PATH',
                        help="path to the table with ID")
    parser.add_argument("-k", required=False, metavar='VALUE_FORMULA',
                        help="What to check: by values or by formulas?")
    return parser.parse_args()


def start(root):
    root.destroy()
    args = get_arguments()
    if all([args.i, args.e, args.o, args.j, args.k]):
        compare_definition(args.i, args.e, args.o, args.j, args.k)
    else:
        setup_gui()
