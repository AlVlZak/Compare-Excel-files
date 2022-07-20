# This is a script to compare two Excel files.
__version__ = "1.0"
import openpyxl as op
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import logging
log = logging.getLogger(__name__)


def find_difference(wb1, wb2, wb_output, sheet, max_rows, max_columns, len_of_sheets_to_check, k, separator, formula, wb1_v, wb2_v):
    max_row = min(max_rows, max(wb1[sheet].max_row, wb2[sheet].max_row))
    max_column = min(max_columns, max(wb1[sheet].max_column, wb2[sheet].max_column))

    df1_to_check, df2_to_check, wb_res, sheet = define_sheet(wb1, wb2, wb_output, sheet, sheet)

    print_status(sheet, k, max_row, max_column, len_of_sheets_to_check)

    rows, cols, df1_to_check, df2_to_check = compare_sheet(df1_to_check, df2_to_check, max_row, max_column)
    if formula:
        df1_to_check, df2_to_check = replace_formulas_by_values(wb1_v[sheet], wb2_v[sheet], max_row, max_column)
    rows = filling_output_df(df1_to_check, df2_to_check, rows, cols, separator)
    add_comparing_values_to_output(wb_res, rows, separator)


def print_status(sheet, k, max_row, max_column, len_of_sheets_to_check):
    status = "Working with sheet \"{}\", max row is {}, max column is {},  progress {}/{}".format(sheet, max_row,
                                                                                                   max_column, k,
                                                                                                   len_of_sheets_to_check)
    print(status)
    log.info(status)


def filling_output_df(df1_to_check, df2_to_check, rows, cols, separator):
    for item in zip(rows, cols):
        output = '{} {} {}'.format(df1_to_check.iloc[item[0], item[1]], separator, df2_to_check.iloc[item[0], item[1]])
        empty_cell = [f"nan {separator} nan", f"None {separator} ", f"nan {separator} ", f" {separator} None", f" {separator} nan"]
        if output in empty_cell:
            output = ""
        output = output.replace("None", "")
        output = output.replace("nan", "")
        df1_to_check.iloc[item[0], item[1]] = output
    rows = dataframe_to_rows(df1_to_check, index=False, header=False)
    return rows


def add_comparing_values_to_output(wb_res, rows, separator):
    color_list = []
    color_tab = True
    color_dict = {"green": "99ffcc", "red": "ff9999", "blue": "33ccff"}
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            wb_res.cell(row=r_idx, column=c_idx, value=value)
            if separator in str(value):
                color_list.append([r_idx, c_idx])
    wb_res.sheet_properties.tabColor = color_dict["green"]
    for c in color_list:
        wb_res.cell(row=c[0], column=c[1]).fill = op.styles.PatternFill("solid", start_color="ff9999")
        if color_tab:
            wb_res.sheet_properties.tabColor = color_dict["red"]
            color_tab = False
    if wb_res.cell(row=1, column=1).value == "" or wb_res.cell(row=1, column=1).value is None:
        wb_res.cell(row=1, column=1, value="Go to sheet Run")
    wb_res.cell(row=1, column=1).hyperlink = "#Run!A1"
    wb_res.cell(row=1, column=1).style = "Hyperlink"


def replace_formulas_by_values(wb1_v, wb2_v, max_row, max_column):
    df1_to_check = pd.DataFrame(wb1_v.values)
    df2_to_check = pd.DataFrame(wb2_v.values)
    df1_to_check, df2_to_check = check_shape_and_set_to_one_shape(df1_to_check, df2_to_check, max_row, max_column)
    return df1_to_check, df2_to_check


def define_sheet(wb1, wb2, wb_output, sheet1, sheet2=""):
    if (sheet1 == sheet2) or (sheet2 == ""):
        sheet = sheet1
        sheet2 = sheet1
    else:
        sheet = sheet1[:15] + "&" + sheet2[:15]
    #wb_output.create_sheet(sheet)
    wb1_to_check = wb1[sheet1]
    wb2_to_check = wb2[sheet2]
    wb_res = wb_output[sheet]
    df1_to_check = pd.DataFrame(wb1_to_check.values)
    df2_to_check = pd.DataFrame(wb2_to_check.values)
    return df1_to_check, df2_to_check, wb_res, sheet


def compare_sheet(df1_to_check, df2_to_check, max_row, max_column):
    df1_to_check, df2_to_check = check_shape_and_set_to_one_shape(df1_to_check, df2_to_check, max_row, max_column)
    comparison_values = df1_to_check.values == df2_to_check.values
    rows, cols = np.where(~comparison_values)
    return rows, cols, df1_to_check, df2_to_check


def check_shape_and_set_to_one_shape(df1_to_check, df2_to_check, max_row, max_column):
    shape1 = df1_to_check.shape
    shape2 = df2_to_check.shape
    if shape1 != shape2:
        if shape1[0] < shape2[0]:
            insert_ = pd.DataFrame([[""] * shape1[1]], columns=df1_to_check.columns)
            diff = shape2[0] - shape1[0]
            for i in range(diff):
                df1_to_check = pd.concat([df1_to_check, insert_], ignore_index=True)
        else:
            insert_ = pd.DataFrame([[""] * shape2[1]], columns=df2_to_check.columns)
            diff = shape1[0] - shape2[0]
            for i in range(diff):
                df2_to_check = pd.concat([df2_to_check, insert_], ignore_index=True)

        if shape1[1] < shape2[1]:
            diff = shape2[1] - shape1[1]
            for i in range(diff):
                df1_to_check[shape1[1] + i] = ""
        else:
            diff = shape1[1] - shape2[1]
            for i in range(diff):
                df2_to_check[shape2[1] + i] = ""
    df1_to_check = df1_to_check.head(max_row)
    df2_to_check = df2_to_check.head(max_row)
    df1_to_check = df1_to_check.iloc[:, :max_column]
    df2_to_check = df2_to_check.iloc[:, :max_column]
    return df1_to_check, df2_to_check